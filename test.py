from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook
from bs4 import BeautifulSoup

def read_tags_from_file(filename):
    with open(filename, 'r', encoding='utf-8') as file:
        tags = [line.strip().split(':') for line in file.readlines()]
    return tags

def scrape_data(html_content, tags):
    soup = BeautifulSoup(html_content, 'html.parser')
    data = {f"{attr}:{value}": set() for attr, value in tags}  # Используем множества для уникальности

    for div in soup.find_all('div'):
        for attr, value in tags:
            if attr == "itemprop" and value == "name":
                span = div.find('span', itemprop=value)
                if span:
                    data[f"{attr}:{value}"].add(span.text)
            elif attr == "itemprop" and value == "url":
                link = div.find('link', itemprop=value)
                if link:
                    data[f"{attr}:{value}"].add(link['href'])
            elif attr == "itemprop" and value == "price":
                meta = div.find('meta', itemprop=value)
                if meta:
                    data[f"{attr}:{value}"].add(meta['content'])
    return {k: list(v) for k, v in data.items()}  # Конвертируем обратно в списки для записи в Excel

# Настройка драйвера и переход на нужную страницу
browser = webdriver.Chrome()
browser.get('https://www.divan.ru/tolyatti/category/svet')
time.sleep(2)  # Даем время для полной загрузки страницы

# Получение HTML-кода страницы
html_content = browser.page_source

# Чтение тегов из файла
tags = read_tags_from_file('tegs.txt')

# Скрапинг данных
data = scrape_data(html_content, tags)

# Создаем новую книгу Excel
wb = Workbook()
ws = wb.active

# Записываем данные в столбцы Excel
for idx, (tag, values) in enumerate(data.items(), 1):
    ws[f'{chr(64 + idx)}1'] = tag  # Записываем имя тега в заголовок столбца
    for row_idx, value in enumerate(values, 2):  # Начинаем с 2-ой строки, так как 1-я строка под заголовки
        ws[f'{chr(64 + idx)}{row_idx}'] = value

# Сохраняем книгу
wb.save("svet.xlsx")

# Закрываем браузер
browser.quit()